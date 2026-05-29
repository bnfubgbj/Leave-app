import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from datetime import date
from io import BytesIO
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

st.set_page_config(page_title="ระบบใบลา", page_icon="🌿", layout="centered")

SHEET_ID = "1KR8adBRkkrEhjgY1lf4eTQ3rovLmcLYJdNCSFN827no"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# ===============================
# Email Notification
# ===============================
def send_email_notification(leave, receiver=None):
    try:
        cfg = st.secrets["email"]
        sender   = cfg["sender"]
        password = cfg["password"].replace(" ", "")
        if not receiver:
            receiver = cfg["receiver"]
        receivers = [r.strip() for r in str(receiver).split(",") if r.strip()]
        if not receivers:
            receivers = [cfg["receiver"]]

        subject = f"[ระบบใบลา] {leave['ชื่อ']} ขอ{leave['ประเภท']} {leave['จำนวนวัน']} วัน รอการอนุมัติ"
        body = f"""
        <h2>📋 มีคำขอลาใหม่รอการอนุมัติ</h2>
        <table border="1" cellpadding="8" style="border-collapse:collapse;">
            <tr><td><b>พนักงาน</b></td><td>{leave['ชื่อ']}</td></tr>
            <tr><td><b>แผนก</b></td><td>{leave['แผนก']}</td></tr>
            <tr><td><b>ตำแหน่ง</b></td><td>{leave['ตำแหน่ง']}</td></tr>
            <tr><td><b>ประเภทการลา</b></td><td>{leave['ประเภท']}</td></tr>
            <tr><td><b>วันที่</b></td><td>{leave['วันเริ่ม']} ถึง {leave['วันสิ้นสุด']}</td></tr>
            <tr><td><b>จำนวนวัน</b></td><td>{leave['จำนวนวัน']} วัน</td></tr>
            <tr><td><b>เหตุผล</b></td><td>{leave['เหตุผล']}</td></tr>
        </table>
        <br>
        <p>กรุณาเข้าระบบเพื่ออนุมัติหรือปฏิเสธคำขอ</p>
        """
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = sender
        msg.attach(MIMEText(body, "html", "utf-8"))
        msg["To"] = ", ".join(receivers)
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender, password)
            server.sendmail(sender, receivers, msg.as_string())
        return True
    except Exception as e:
        st.warning(f"ส่ง Email ไม่สำเร็จ: {e}")
        return False

# ===============================
# Google Sheets Connection
# ===============================
@st.cache_resource(ttl=60)
def get_client():
    creds_dict = st.secrets["gcp_service_account"]
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return gspread.authorize(creds)

def get_sheet(name):
    client = get_client()
    return client.open_by_key(SHEET_ID).worksheet(name)

# ===============================
# Employees
# ===============================
def load_employees(include_inactive=False):
    try:
        ws = get_sheet("employees")
        data = ws.get_all_records(numericise_ignore=["all"])
        for e in data:
            if "รหัส" in e:
                e["รหัส"] = str(e["รหัส"]).strip().zfill(4)
        if not include_inactive:
            data = [e for e in data if str(e.get("สถานะ","active")).strip().lower() != "inactive"]
        return data
    except:
        return []

def save_employee(emp):
    ws = get_sheet("employees")
    records = ws.get_all_values()
    if len(records) == 0:
        ws.append_row(["รหัส","ชื่อ","ตำแหน่ง","แผนก","วันเริ่มงาน","ลาพักร้อน","ลากิจ","ลาป่วย"])
    ws.append_row([
        emp["รหัส"], emp["ชื่อ"], emp["ตำแหน่ง"], emp["แผนก"],
        emp["วันเริ่มงาน"], emp["ลาพักร้อน"], emp["ลากิจ"], emp["ลาป่วย"],
        emp.get("รหัสหัวหน้า",""), emp.get("อีเมลหัวหน้า",""),
        emp.get("วันสะสม","0"), emp.get("รหัสผ่าน",""), "active"
    ])

def update_employee(emp_id, updated):
    ws = get_sheet("employees")
    records = ws.get_all_values()
    for i, row in enumerate(records):
        if row[0] == emp_id:
            ws.update(f"A{i+1}:H{i+1}", [[
                emp_id, updated["ชื่อ"], updated["ตำแหน่ง"], updated["แผนก"],
                row[4], updated["ลาพักร้อน"], updated["ลากิจ"], updated["ลาป่วย"]
            ]])
            break

def delete_employee(emp_id):
    ws = get_sheet("employees")
    records = ws.get_all_values()
    for i, row in enumerate(records):
        if row[0] == emp_id:
            ws.delete_rows(i + 1)
            break

def normalize_id(emp_id):
    return str(emp_id).strip().zfill(4)

def get_employee(emp_id):
    nid = normalize_id(emp_id)
    for e in load_employees():
        if normalize_id(str(e.get("รหัส",""))) == nid:
            return e
    return None

# ===============================
# Leaves
# ===============================
def load_leaves():
    try:
        ws = get_sheet("leaves")
        data = ws.get_all_records(numericise_ignore=["all"])
        return data
    except:
        return []

def save_leave(leave):
    ws = get_sheet("leaves")
    records = ws.get_all_values()
    if len(records) == 0:
        ws.append_row(["id","รหัส","ชื่อ","แผนก","ตำแหน่ง","ประเภท","วันเริ่ม","วันสิ้นสุด","จำนวนวัน","เหตุผล","ผู้อนุมัติ","สถานะ","หมายเหตุ"])
    ws.append_row([
        leave["id"], leave["รหัส"], leave["ชื่อ"], leave["แผนก"], leave["ตำแหน่ง"],
        leave["ประเภท"], leave["วันเริ่ม"], leave["วันสิ้นสุด"], leave["จำนวนวัน"],
        leave["เหตุผล"], leave["ผู้อนุมัติ"], leave["สถานะ"], leave["หมายเหตุ"]
    ])

def update_leave_status(leave_id, status, note):
    ws = get_sheet("leaves")
    records = ws.get_all_values()
    for i, row in enumerate(records):
        if i == 0:
            continue
        if str(row[0]).strip() == str(leave_id).strip():
            ws.update_cell(i + 1, 12, status)
            ws.update_cell(i + 1, 13, note)
            break

def deactivate_employee(emp_id):
    ws = get_sheet("employees")
    records = ws.get_all_values()
    headers = records[0] if records else []
    try:
        status_col = headers.index("สถานะ") + 1
    except:
        status_col = len(headers) + 1
        ws.update_cell(1, status_col, "สถานะ")
    for i, row in enumerate(records):
        if i == 0:
            continue
        if str(row[0]).strip().zfill(4) == emp_id:
            ws.update_cell(i + 1, status_col, "inactive")
            break

def activate_employee(emp_id):
    ws = get_sheet("employees")
    records = ws.get_all_values()
    headers = records[0] if records else []
    try:
        status_col = headers.index("สถานะ") + 1
    except:
        return
    for i, row in enumerate(records):
        if i == 0:
            continue
        if str(row[0]).strip().zfill(4) == emp_id:
            ws.update_cell(i + 1, status_col, "active")
            break

def get_used_days(emp_id, leave_type=None):
    leaves = load_leaves()
    result = [l for l in leaves if str(l.get("รหัส","")) == str(emp_id)
              and l.get("สถานะ","") in ["อนุมัติแล้ว", "รออนุมัติ"]]
    if leave_type:
        result = [l for l in result if l.get("ประเภท","") == leave_type]
    return sum(int(l.get("จำนวนวัน",0)) for l in result)

def get_approved_days(emp_id, leave_type=None):
    leaves = load_leaves()
    result = [l for l in leaves if str(l.get("รหัส","")) == str(emp_id)
              and l.get("สถานะ","") == "อนุมัติแล้ว"]
    if leave_type:
        result = [l for l in result if l.get("ประเภท","") == leave_type]
    return sum(int(l.get("จำนวนวัน",0)) for l in result)

def get_pending_days(emp_id, leave_type=None):
    leaves = load_leaves()
    result = [l for l in leaves if str(l.get("รหัส","")) == str(emp_id)
              and l.get("สถานะ","") == "รออนุมัติ"]
    if leave_type:
        result = [l for l in result if l.get("ประเภท","") == leave_type]
    return sum(int(l.get("จำนวนวัน",0)) for l in result)

def check_overlap(emp_id, start, end):
    from datetime import datetime
    leaves = load_leaves()
    pending = [l for l in leaves if str(l.get("รหัส","")) == str(emp_id)
               and l.get("สถานะ","") in ["รออนุมัติ", "อนุมัติแล้ว"]]
    new_start = datetime.strptime(str(start), "%Y-%m-%d")
    new_end   = datetime.strptime(str(end), "%Y-%m-%d")
    overlaps  = []
    for l in pending:
        try:
            l_start = datetime.strptime(str(l.get("วันเริ่ม","")), "%Y-%m-%d")
            l_end   = datetime.strptime(str(l.get("วันสิ้นสุด","")), "%Y-%m-%d")
            if new_start <= l_end and new_end >= l_start:
                overlaps.append(l)
        except:
            pass
    return overlaps

# ===============================
# Config
# ===============================
def load_config():
    try:
        ws = get_sheet("config")
        data = ws.get_all_records()
        config = {row["key"]: row["value"] for row in data}
        if not config:
            raise Exception("empty")
        return config
    except:
        return {
            "admin_password": "admin1234",
            "secret_question": "ชื่อบริษัทของคุณคืออะไร?",
            "secret_answer": "abc"
        }

def save_config(config):
    ws = get_sheet("config")
    ws.clear()
    ws.append_row(["key", "value"])
    for k, v in config.items():
        ws.append_row([k, v])

# ===============================
# Excel Export
# ===============================
def export_summary_excel(employees):
    rows = []
    for e in employees:
        used_annual   = get_used_days(str(e.get("รหัส","")), "ลาพักร้อน")
        used_personal = get_used_days(e["รหัส"], "ลากิจ")
        used_sick     = get_used_days(e["รหัส"], "ลาป่วย")
        rows.append({
            "รหัส": e["รหัส"], "ชื่อ-นามสกุล": e["ชื่อ"],
            "ตำแหน่ง": e["ตำแหน่ง"], "แผนก": e["แผนก"],
            "สิทธิ์ลาพักร้อน": e["ลาพักร้อน"],
            "ลาพักร้อนที่ใช้": used_annual,
            "ลาพักร้อนคงเหลือ": int(e["ลาพักร้อน"]) - used_annual,
            "สิทธิ์ลากิจ": e["ลากิจ"],
            "ลากิจที่ใช้": used_personal,
            "ลากิจคงเหลือ": int(e["ลากิจ"]) - used_personal,
            "สิทธิ์ลาป่วย": e["ลาป่วย"],
            "ลาป่วยที่ใช้": used_sick,
            "ลาป่วยคงเหลือ": int(e["ลาป่วย"]) - used_sick,
        })
    df = pd.DataFrame(rows)
    num_cols = ["สิทธิ์ลาพักร้อน","ลาพักร้อนที่ใช้","ลาพักร้อนคงเหลือ",
                "สิทธิ์ลากิจ","ลากิจที่ใช้","ลากิจคงเหลือ",
                "สิทธิ์ลาป่วย","ลาป่วยที่ใช้","ลาป่วยคงเหลือ"]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="สรุปการลา")
        ws = writer.sheets["สรุปการลา"]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        row_fill1 = PatternFill("solid", fgColor="D6E4F0")
        row_fill2 = PatternFill("solid", fgColor="FFFFFF")
        thin = Side(style="thin", color="B8CCE4")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = row_fill1 if row % 2 == 0 else row_fill2
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 25)
        ws.row_dimensions[1].height = 35
    return output.getvalue()

# ===============================
# Login System
# ===============================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.login_user = None
    st.session_state.is_admin = False

def do_logout():
    st.session_state.logged_in = False
    st.session_state.login_user = None
    st.session_state.is_admin = False

if not st.session_state.logged_in:
    import base64
    try:
        with open("logo.jpg", "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode()
        st.markdown(f"""
        <div style="display:flex; align-items:center; gap:20px; margin-bottom:10px;">
            <img src="data:image/jpeg;base64,{logo_b64}" style="width:80px; object-fit:contain;">
            <div>
                <div style="font-size:2rem; font-weight:700; line-height:1.2;">ระบบจัดการใบลา</div>
                <div style="color:gray; font-size:13px;">VIRTUARCH CO.,LTD</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    except:
        st.markdown("<h1>🌿 ระบบจัดการใบลา</h1>", unsafe_allow_html=True)
        st.markdown("<p style='color:gray;'>VIRTUARCH CO.,LTD</p>", unsafe_allow_html=True)
    st.divider()
    st.subheader("🔐 เข้าสู่ระบบ")
    all_emps = load_employees()
    if not all_emps:
        st.warning("⚠️ กำลังโหลดข้อมูล กรุณารอสักครู่แล้ว refresh หน้าใหม่")
        st.stop()

    with st.form("login_form"):
        emp_options = ["-- เลือกชื่อพนักงาน --"] + [f"{e.get('รหัส','')} - {e.get('ชื่อ','')}" for e in all_emps]
        selected_emp = st.selectbox("เลือกชื่อพนักงาน", emp_options)
        login_pw = st.text_input("รหัสผ่าน", type="password")
        if st.form_submit_button("เข้าสู่ระบบ"):
            config = load_config()
            if selected_emp == "-- เลือกชื่อพนักงาน --":
                st.error("❌ กรุณาเลือกชื่อพนักงาน")
            else:
                login_id = selected_emp.split(" - ")[0].strip()
                if login_id == "admin" and login_pw == config.get("admin_password", "admin1234"):
                    st.session_state.logged_in = True
                    st.session_state.login_user = None
                    st.session_state.is_admin = True
                    st.rerun()
                else:
                    emp = get_employee(login_id)
                    if emp and str(emp.get("รหัสผ่าน","")).strip() == str(login_pw).strip():
                        st.session_state.logged_in = True
                        st.session_state.login_user = emp
                        st.session_state.is_admin = False
                        st.rerun()
                    else:
                        st.error("❌ รหัสผ่านไม่ถูกต้อง")

    st.divider()
    st.caption("สำหรับ Admin")
    with st.form("admin_login_form"):
        admin_pw = st.text_input("รหัสผ่าน Admin", type="password")
        if st.form_submit_button("เข้าสู่ระบบ Admin"):
            config = load_config()
            if admin_pw == config.get("admin_password", "admin1234"):
                st.session_state.logged_in = True
                st.session_state.login_user = None
                st.session_state.is_admin = True
                st.rerun()
            else:
                st.error("❌ รหัสผ่าน Admin ไม่ถูกต้อง")
    st.stop()

# ===============================
# Post-login
# ===============================
current_user = st.session_state.login_user
is_admin = st.session_state.is_admin

with st.sidebar:
    try:
        st.image("logo.jpg", width=120)
    except:
        st.markdown("### 🌿 VIRTUARCH")
    st.caption("VIRTUARCH CO.,LTD | ระบบจัดการใบลา")
    st.divider()
    if is_admin:
        st.success("👑 Admin")
    else:
        st.success(f"👤 {current_user.get('ชื่อ','')} ({current_user.get('รหัส','')})")
    if st.button("🚪 ออกจากระบบ"):
        do_logout()
        st.rerun()

if is_admin:
    menu = st.sidebar.radio("เมนู", [
        "📋 ประวัติการลา",
        "👥 จัดการพนักงาน (Admin)",
    ])
else:
    my_id = current_user.get("รหัส","") if current_user else ""
    all_leaves = load_leaves()
    all_emps_sidebar = load_employees()

    subordinates_sidebar = [e.get("รหัส","") for e in all_emps_sidebar
                           if normalize_id(str(e.get("รหัสหัวหน้า",""))) == normalize_id(my_id)]

    def is_for_me(l):
        leave_emp_id = normalize_id(str(l.get("รหัส","")))
        approver_field = str(l.get("ผู้อนุมัติ",""))
        approver_name = current_user.get("ชื่อ","") if current_user else ""
        is_sub = leave_emp_id in [normalize_id(s) for s in subordinates_sidebar]
        is_named = my_id in approver_field or approver_name in approver_field
        return is_sub or is_named

    waiting_for_me = [l for l in all_leaves
                     if l.get("สถานะ","") == "รออนุมัติ"
                     and normalize_id(str(l.get("รหัส",""))) != normalize_id(my_id)
                     and is_for_me(l)]

    my_pending = [l for l in all_leaves
                 if l.get("สถานะ","") == "รออนุมัติ"
                 and normalize_id(str(l.get("รหัส",""))) == normalize_id(my_id)]

    if waiting_for_me:
        st.sidebar.error(f"📋 {len(waiting_for_me)} รายการรอการอนุมัติจากคุณ")
    if my_pending:
        st.sidebar.warning(f"⏳ {len(my_pending)} รายการของคุณรออนุมัติอยู่")

    approve_label = f"✅ อนุมัติใบลา ({len(waiting_for_me)} ค้าง)" if waiting_for_me else "✅ อนุมัติใบลา"
    menu = st.sidebar.radio("เมนู", [
        "📝 ยื่นคำขอลา",
        approve_label,
        "📋 ประวัติการลา",
    ])
    if "อนุมัติใบลา" in menu:
        menu = "✅ อนุมัติใบลา"

# ===============================
# หน้า: ยื่นคำขอลา
# ===============================
if menu == "📝 ยื่นคำขอลา":
    st.header("ยื่นคำขอลา")
    emp = current_user
    emp_id = emp.get("รหัส","") if emp else ""
    if emp:
        carry         = int(emp.get("วันสะสม", 0) or 0)
        total_annual  = int(emp.get("ลาพักร้อน", 0)) + carry
        approved_annual   = get_approved_days(emp["รหัส"], "ลาพักร้อน")
        pending_annual    = get_pending_days(emp["รหัส"], "ลาพักร้อน")
        approved_personal = get_approved_days(emp["รหัส"], "ลากิจ")
        pending_personal  = get_pending_days(emp["รหัส"], "ลากิจ")
        approved_sick     = get_approved_days(emp["รหัส"], "ลาป่วย")
        pending_sick      = get_pending_days(emp["รหัส"], "ลาป่วย")
        left_annual   = total_annual - approved_annual - pending_annual
        left_personal = int(emp.get("ลากิจ", 0)) - approved_personal - pending_personal
        left_sick     = int(emp.get("ลาป่วย", 0)) - approved_sick - pending_sick

        st.success(f"👤 {emp['ชื่อ']} — {emp['ตำแหน่ง']} ({emp['แผนก']})")

        df_quota = pd.DataFrame([
            {"ประเภท": "🏖 ลาพักร้อน", "สิทธิ์รวม": total_annual,
             "อนุมัติแล้ว": approved_annual, "รออนุมัติ": pending_annual, "คงเหลือ": left_annual},
            {"ประเภท": "📋 ลากิจ", "สิทธิ์รวม": int(emp.get("ลากิจ", 0)),
             "อนุมัติแล้ว": approved_personal, "รออนุมัติ": pending_personal, "คงเหลือ": left_personal},
            {"ประเภท": "🏥 ลาป่วย", "สิทธิ์รวม": int(emp.get("ลาป่วย", 0)),
             "อนุมัติแล้ว": approved_sick, "รออนุมัติ": pending_sick, "คงเหลือ": left_sick},
        ])
        st.dataframe(df_quota, use_container_width=True, hide_index=True)
        if left_annual <= 0 or left_personal <= 0 or left_sick <= 0:
            st.error("⚠️ วันลาบางประเภทหมดแล้ว!")

        all_leaves_emp = load_leaves()
        pending_leaves = [l for l in all_leaves_emp
                        if normalize_id(str(l.get("รหัส",""))) == normalize_id(emp_id)
                        and l.get("สถานะ","") == "รออนุมัติ"]
        if pending_leaves:
            st.warning("⏳ รายการที่รออนุมัติ")
            df_pending_show = pd.DataFrame([{
                "ประเภท": pl.get("ประเภท",""),
                "วันเริ่ม": pl.get("วันเริ่ม",""),
                "วันสิ้นสุด": pl.get("วันสิ้นสุด",""),
                "จำนวนวัน": pl.get("จำนวนวัน",""),
                "เหตุผล": pl.get("เหตุผล",""),
            } for pl in pending_leaves])
            st.dataframe(df_pending_show, use_container_width=True, hide_index=True)

    if emp:
        from datetime import datetime as dt2, timedelta as td2

        # โหลดวันที่ลาไปแล้ว
        existing_leaves = load_leaves()
        booked_dates_set = set()
        booked_info = {}
        for l in existing_leaves:
            if (normalize_id(str(l.get("รหัส",""))) == normalize_id(emp["รหัส"])
                    and l.get("สถานะ","") in ["รออนุมัติ","อนุมัติแล้ว"]):
                try:
                    s2 = dt2.strptime(str(l.get("วันเริ่ม","")), "%Y-%m-%d").date()
                    e2 = dt2.strptime(str(l.get("วันสิ้นสุด","")), "%Y-%m-%d").date()
                    short = {"ลาพักร้อน":"พักร้อน","ลาป่วย":"ป่วย","ลากิจ":"กิจ","อื่นๆ":"อื่นๆ"}
                    label = short.get(l.get("ประเภท",""), l.get("ประเภท",""))
                    cur2 = s2
                    while cur2 <= e2:
                        booked_dates_set.add(cur2)
                        booked_info[str(cur2)] = label
                        cur2 += td2(days=1)
                except:
                    pass

        # แสดงวันที่ลาไปแล้ว แบบ Badge จัดกลุ่มตามเดือน
        if booked_info:
            booked_list = sorted(booked_info.keys())
            total_booked = len(booked_list)

            color_map = {
                "พักร้อน": "#1565C0",
                "ป่วย":    "#B71C1C",
                "กิจ":     "#1B5E20",
                "อื่นๆ":   "#4A148C",
            }
            th_month = ["","มกราคม","กุมภาพันธ์","มีนาคม","เมษายน","พฤษภาคม",
                        "มิถุนายน","กรกฎาคม","สิงหาคม","กันยายน","ตุลาคม",
                        "พฤศจิกายน","ธันวาคม"]

            # จัดกลุ่มตามเดือน
            from collections import defaultdict
            grouped = defaultdict(list)
            for d in booked_list:
                ym = d[:7]  # "2026-06"
                grouped[ym].append(d)

            # สร้าง HTML
            html_parts = []
            for ym in sorted(grouped.keys()):
                y, m = ym.split("-")
                month_label = f"{th_month[int(m)]} พ.ศ.{int(y)+543}"
                badges = ""
                for d in grouped[ym]:
                    day_num = d.split("-")[2]
                    label = booked_info[d]
                    color = color_map.get(label, "#546E7A")
                    badges += f"""<span style="
                        background:{color};color:white;
                        padding:3px 10px;border-radius:20px;
                        font-size:12px;margin:3px 4px 3px 0;
                        display:inline-block;font-family:sans-serif;">
                        {day_num} {th_month[int(m)][:3]}. &nbsp;{label}
                    </span>"""
                html_parts.append(f"""
                    <div style="margin-bottom:8px;">
                        <div style="font-size:12px;color:#888;margin-bottom:4px;">{month_label}</div>
                        <div>{badges}</div>
                    </div>
                """)

            with st.expander(f"📅 วันที่ลาไปแล้ว ({total_booked} วัน)"):
                import streamlit.components.v1 as _components
                _components.html(
                    "<div style='padding:4px 0;font-family:sans-serif;'>" + "".join(html_parts) + "</div>",
                    height=min(total_booked * 20 + 80, 250),
                    scrolling=False
                )

        # เลือกวันที่ลา — ช่วงเดียว
        st.markdown("**📅 เลือกวันที่ต้องการลา**")
        col_s, col_e = st.columns(2)
        with col_s:
            leave_start = st.date_input("วันเริ่มต้น", value=date.today(), key="leave_start")
        with col_e:
            leave_end = st.date_input("วันสิ้นสุด", value=date.today(), key="leave_end")

        selected_dates = []
        overlap_dates  = []
        if leave_start <= leave_end:
            cur = leave_start
            while cur <= leave_end:
                if cur in booked_dates_set:
                    overlap_dates.append(str(cur))
                else:
                    selected_dates.append(str(cur))
                cur += td2(days=1)

        total_days = len(selected_dates)
        has_overlap = len(overlap_dates) > 0

        if leave_start > leave_end:
            st.error("❌ วันเริ่มต้นต้องไม่มากกว่าวันสิ้นสุด")
        elif has_overlap:
            st.error(f"⚠️ มีวันที่ซ้ำกับวันลาที่มีอยู่แล้ว: {', '.join(overlap_dates)}")
            if selected_dates:
                st.warning(f"วันที่ใช้ได้: {', '.join(selected_dates)} | รวม {total_days} วัน")
        elif total_days > 0:
            st.success(f"✅ เลือก **{total_days} วัน**: {', '.join(selected_dates)}")

        start = leave_start
        end   = leave_end
        days  = total_days

        # selectbox นอก form เพื่อให้ rerun ทันทีเมื่อเปลี่ยนประเภท
        quota_map = {
            "ลาพักร้อน": left_annual,
            "ลากิจ":     left_personal,
            "ลาป่วย":    left_sick,
            "อื่นๆ":     999
        }
        leave_type = st.selectbox("ประเภทการลา", ["ลาพักร้อน", "ลาป่วย", "ลากิจ", "อื่นๆ"], key="leave_type_sel")
        remaining  = quota_map.get(leave_type, 0)
        rem_icon   = "🟢" if remaining > 3 else ("🟡" if remaining > 0 else "🔴")
        st.caption(f"{rem_icon} สิทธิ์คงเหลือ: **{remaining} วัน**")

        with st.form("leave_form"):
            st.markdown(f"**สรุป:** {leave_type} | สิทธิ์คงเหลือ {remaining} วัน")
            reason = st.text_area("เหตุผล")
            boss_id_val = str(emp.get("รหัสหัวหน้า", "")).strip().zfill(4)
            boss_emp = get_employee(boss_id_val) if boss_id_val and boss_id_val != "0000" else None
            if boss_emp:
                approver = f"{boss_id_val} - {boss_emp.get('ชื่อ','')} ({boss_emp.get('ตำแหน่ง','')})"
                st.info(f"👤 ผู้อนุมัติ: **{approver}**")
            else:
                approver = st.selectbox("ผู้อนุมัติ", ["ผู้จัดการ", "HR"])
            submitted = st.form_submit_button("📤 ส่งคำขอลา")
            if submitted:
                if days <= 0 or not selected_dates:
                    st.error("⚠️ กรุณาเลือกวันที่ต้องการลาในปฏิทินก่อนครับ")
                elif has_overlap:
                    st.error("❌ กรุณาแก้ไขช่วงวันที่ซ้ำกับวันลาที่มีอยู่แล้ว")
                elif leave_type == "ลาพักร้อน" and days > left_annual:
                    st.error(f"❌ วันลาพักร้อนไม่เพียงพอ! คงเหลือ {left_annual} วัน แต่ขอลา {days} วัน")
                elif leave_type == "ลากิจ" and days > left_personal:
                    st.error(f"❌ วันลากิจไม่เพียงพอ! คงเหลือ {left_personal} วัน แต่ขอลา {days} วัน")
                elif leave_type == "ลาป่วย" and days > left_sick:
                    st.error(f"❌ วันลาป่วยไม่เพียงพอ! คงเหลือ {left_sick} วัน แต่ขอลา {days} วัน")
                elif not reason:
                    st.error("กรุณาระบุเหตุผล")
                else:
                    leaves = load_leaves()
                    new_leave = {
                        "id": len(leaves) + 1,
                        "รหัส": emp["รหัส"], "ชื่อ": emp["ชื่อ"],
                        "แผนก": emp["แผนก"], "ตำแหน่ง": emp["ตำแหน่ง"],
                        "ประเภท": leave_type, "วันเริ่ม": str(start),
                        "วันสิ้นสุด": str(end), "จำนวนวัน": days,
                        "เหตุผล": reason, "ผู้อนุมัติ": approver,
                        "สถานะ": "รออนุมัติ", "หมายเหตุ": ""
                    }
                    save_leave(new_leave)
                    boss_email = emp.get("อีเมลหัวหน้า", "")
                    send_email_notification(new_leave, receiver=boss_email if boss_email else None)
                    st.success("✅ ยื่นคำขอลาเรียบร้อยแล้ว! แจ้งเตือนผู้บังคับบัญชาแล้ว 📧")

# ===============================
# หน้า: อนุมัติใบลา
# ===============================
elif menu == "✅ อนุมัติใบลา":
    st.header("อนุมัติใบลา")
    approver_emp = current_user
    approver_id = approver_emp.get("รหัส","") if approver_emp else ""

    if waiting_for_me:
        st.error(f"📋 มี **{len(waiting_for_me)} รายการ** รอการอนุมัติจากคุณ")
    if my_pending:
        st.warning(f"⏳ ใบลาของคุณ **{len(my_pending)} รายการ** กำลังรออนุมัติอยู่")
    st.divider()

    leaves = load_leaves()
    all_emps = load_employees()
    subordinates = [e.get("รหัส","") for e in all_emps
                   if normalize_id(str(e.get("รหัสหัวหน้า",""))) == normalize_id(approver_id)]

    def is_for_approver(l):
        leave_emp_id = normalize_id(str(l.get("รหัส","")))
        approver_field = str(l.get("ผู้อนุมัติ",""))
        is_subordinate = leave_emp_id in [normalize_id(s) for s in subordinates]
        approver_name = current_user.get("ชื่อ","") if current_user else ""
        is_named = approver_id in approver_field or approver_name in approver_field
        return is_subordinate or is_named

    pending = [l for l in leaves if l.get("สถานะ","") == "รออนุมัติ"
               and normalize_id(str(l.get("รหัส",""))) != normalize_id(approver_id)
               and is_for_approver(l)]

    if not pending:
        st.info("✅ ไม่มีรายการรออนุมัติในขณะนี้")
    for i, l in enumerate(pending):
        with st.expander(f"📋 {l['ชื่อ']} — {l['ประเภท']} ({l['จำนวนวัน']} วัน)"):
            st.write(f"**แผนก:** {l['แผนก']} | **ตำแหน่ง:** {l['ตำแหน่ง']}")
            st.write(f"**วันที่:** {l['วันเริ่ม']} ถึง {l['วันสิ้นสุด']}")
            st.write(f"**เหตุผล:** {l['เหตุผล']}")
            note = st.text_input("หมายเหตุ", key=f"note_{i}")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("✅ อนุมัติ", key=f"approve_{i}"):
                    update_leave_status(l.get("id", l.get("ID", l.get("Id",""))), "อนุมัติแล้ว", note)
                    st.success("อนุมัติเรียบร้อย!")
                    st.rerun()
            with col2:
                if st.button("❌ ปฏิเสธ", key=f"reject_{i}"):
                    update_leave_status(l.get("id", l.get("ID", l.get("Id",""))), "ถูกปฏิเสธ", note)
                    st.error("ปฏิเสธแล้ว")
                    st.rerun()

# ===============================
# หน้า: ประวัติการลา
# ===============================
elif menu == "📋 ประวัติการลา":
    st.header("ประวัติการลา")
    leaves = load_leaves()
    all_emps = load_employees()

    if not leaves:
        st.info("ยังไม่มีประวัติการลา")
    else:
        df = pd.DataFrame(leaves)
        df = df[["รหัส","ชื่อ","แผนก","ประเภท","วันเริ่ม","วันสิ้นสุด","จำนวนวัน","สถานะ","เหตุผล"]]

        if is_admin:
            emp_filter = st.selectbox("กรองตามพนักงาน", ["ทั้งหมด"] + [f"{e['รหัส']} - {e['ชื่อ']}" for e in all_emps])
            if emp_filter != "ทั้งหมด":
                filter_id = emp_filter.split(" - ")[0]
                df = df[df["รหัส"] == filter_id]
        else:
            my_id = current_user.get("รหัส","")
            subordinates = [e.get("รหัส","") for e in all_emps
                          if normalize_id(str(e.get("รหัสหัวหน้า",""))) == normalize_id(my_id)]
            if subordinates:
                allowed_ids = [my_id] + subordinates
                df = df[df["รหัส"].apply(lambda x: normalize_id(str(x)) in [normalize_id(i) for i in allowed_ids])]
                view_options = ["ของฉัน"] + [f"{e['รหัส']} - {e['ชื่อ']}" for e in all_emps if e.get("รหัส","") in subordinates]
                selected_view = st.selectbox("ดูประวัติของ", ["ทั้งหมด (ฉัน + ลูกน้อง)"] + view_options)
                if selected_view == "ของฉัน":
                    df = df[df["รหัส"].apply(lambda x: normalize_id(str(x)) == normalize_id(my_id))]
                elif selected_view != "ทั้งหมด (ฉัน + ลูกน้อง)":
                    sel_id = selected_view.split(" - ")[0]
                    df = df[df["รหัส"].apply(lambda x: normalize_id(str(x)) == normalize_id(sel_id))]
            else:
                df = df[df["รหัส"].apply(lambda x: normalize_id(str(x)) == normalize_id(my_id))]

        status_filter = st.selectbox("กรองตามสถานะ", ["ทั้งหมด","รออนุมัติ","อนุมัติแล้ว","ถูกปฏิเสธ"])
        if status_filter != "ทั้งหมด":
            df = df[df["สถานะ"] == status_filter]

        if df.empty:
            st.info("ไม่พบรายการ")
        else:
            st.dataframe(df, use_container_width=True, hide_index=True)

# ===============================
# หน้า: จัดการพนักงาน (Admin)
# ===============================
elif menu == "👥 จัดการพนักงาน (Admin)":
    st.header("👥 จัดการพนักงาน")
    if not is_admin:
        st.error("❌ คุณไม่มีสิทธิ์เข้าถึงหน้านี้")
        st.stop()

    employees = load_employees()
    if employees:
        st.subheader("รายชื่อพนักงานในระบบ")
        st.dataframe(pd.DataFrame(employees), use_container_width=True, hide_index=True)
    else:
        st.info("ยังไม่มีพนักงานในระบบ")

    if employees:
        st.divider()
        st.subheader("📊 สรุปการลา")
        st.download_button(
            label="📥 ดาวน์โหลดสรุปการลา (.xlsx)",
            data=export_summary_excel(employees),
            file_name=f"สรุปการลา_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.divider()
    st.subheader("➕ เพิ่มพนักงานใหม่")
    with st.form("add_employee"):
        col1, col2 = st.columns(2)
        with col1:
            emp_id   = st.text_input("รหัสพนักงาน เช่น 0001")
            fullname = st.text_input("ชื่อ-นามสกุล")
            dept     = st.text_input("แผนก")
        with col2:
            position   = st.text_input("ตำแหน่ง")
            start_date = st.date_input("วันเริ่มงาน")
            annual     = st.number_input("สิทธิ์ลาพักร้อน (วัน)", min_value=0, value=6)
            personal   = st.number_input("สิทธิ์ลากิจ (วัน)", min_value=0, value=3)
            sick       = st.number_input("สิทธิ์ลาป่วย (วัน)", min_value=0, value=30)
        boss_id          = st.text_input("รหัสหัวหน้า (ถ้ามี)")
        boss_email_input = st.text_input("อีเมลหัวหน้า (สำหรับแจ้งเตือน)")
        if st.form_submit_button("➕ เพิ่มพนักงาน"):
            if not emp_id or not fullname:
                st.error("กรุณากรอกรหัสและชื่อ-นามสกุล")
            elif get_employee(emp_id.strip()):
                st.error("❌ รหัสพนักงานนี้มีอยู่แล้ว")
            else:
                save_employee({
                    "รหัส": emp_id.strip().zfill(4), "ชื่อ": fullname.strip(),
                    "ตำแหน่ง": position, "แผนก": dept,
                    "วันเริ่มงาน": str(start_date),
                    "ลาพักร้อน": annual, "ลากิจ": personal, "ลาป่วย": sick,
                    "รหัสหัวหน้า": boss_id.strip().zfill(4) if boss_id.strip() else "",
                    "อีเมลหัวหน้า": boss_email_input.strip(),
                })
                st.success(f"✅ เพิ่ม {fullname} เรียบร้อยแล้ว!")
                st.rerun()

    if employees:
        st.divider()
        st.subheader("✏️ แก้ไขข้อมูลพนักงาน")
        emp_list = [f"{e['รหัส']} — {e['ชื่อ']}" for e in employees]
        selected_edit = st.selectbox("เลือกพนักงาน", emp_list, key="edit_select")
        edit_id = selected_edit.split(" — ")[0]
        edit_emp = get_employee(edit_id)
        if edit_emp:
            with st.form("edit_employee"):
                col1, col2 = st.columns(2)
                with col1:
                    new_name = st.text_input("ชื่อ-นามสกุล", value=edit_emp["ชื่อ"])
                    new_dept = st.text_input("แผนก", value=edit_emp["แผนก"])
                with col2:
                    new_position = st.text_input("ตำแหน่ง", value=edit_emp["ตำแหน่ง"])
                    new_annual   = st.number_input("สิทธิ์ลาพักร้อน", min_value=0, value=int(edit_emp["ลาพักร้อน"]))
                    new_personal = st.number_input("สิทธิ์ลากิจ", min_value=0, value=int(edit_emp["ลากิจ"]))
                    new_sick     = st.number_input("สิทธิ์ลาป่วย", min_value=0, value=int(edit_emp["ลาป่วย"]))
                if st.form_submit_button("💾 บันทึก"):
                    update_employee(edit_id, {
                        "ชื่อ": new_name, "แผนก": new_dept, "ตำแหน่ง": new_position,
                        "ลาพักร้อน": new_annual, "ลากิจ": new_personal, "ลาป่วย": new_sick,
                    })
                    st.success("✅ บันทึกเรียบร้อย!")
                    st.rerun()

    if employees:
        st.divider()
        st.subheader("🔴 ปิดการใช้งานพนักงาน")
        st.caption("ข้อมูลยังอยู่ แต่พนักงานจะ Login ไม่ได้")
        all_emps_admin = load_employees(include_inactive=True)
        emp_list2 = [f"{e['รหัส']} — {e['ชื่อ']} {'🔴' if str(e.get('สถานะ','')).lower() == 'inactive' else '🟢'}" for e in all_emps_admin]
        selected_del = st.selectbox("เลือกพนักงาน", emp_list2, key="del_select")
        del_id = selected_del.split(" — ")[0]
        del_emp = next((e for e in all_emps_admin if e.get("รหัส","") == del_id), None)
        col1, col2 = st.columns(2)
        with col1:
            if del_emp and str(del_emp.get("สถานะ","")).lower() != "inactive":
                if st.button("🔴 ปิดการใช้งาน"):
                    deactivate_employee(del_id)
                    st.success(f"ปิดการใช้งาน {del_emp.get('ชื่อ','')} แล้ว!")
                    st.rerun()
        with col2:
            if del_emp and str(del_emp.get("สถานะ","")).lower() == "inactive":
                if st.button("🟢 เปิดการใช้งาน"):
                    activate_employee(del_id)
                    st.success(f"เปิดการใช้งาน {del_emp.get('ชื่อ','')} แล้ว!")
                    st.rerun()

    st.divider()
    st.subheader("🔑 เปลี่ยนรหัสผ่าน")
    with st.form("change_password"):
        old_pw  = st.text_input("รหัสผ่านเดิม", type="password")
        new_pw  = st.text_input("รหัสผ่านใหม่", type="password")
        new_pw2 = st.text_input("ยืนยันรหัสผ่านใหม่", type="password")
        if st.form_submit_button("🔑 เปลี่ยนรหัสผ่าน"):
            config = load_config()
            if old_pw != config["admin_password"]:
                st.error("❌ รหัสผ่านเดิมไม่ถูกต้อง")
            elif new_pw != new_pw2:
                st.error("❌ รหัสผ่านใหม่ไม่ตรงกัน")
            elif len(new_pw) < 6:
                st.error("❌ รหัสผ่านต้องมีอย่างน้อย 6 ตัว")
            else:
                config["admin_password"] = new_pw
                save_config(config)
                st.success("✅ เปลี่ยนรหัสผ่านเรียบร้อย!")

    st.divider()
    st.subheader("❓ ตั้งคำถามลับ")
    config = load_config()
    with st.form("change_secret"):
        new_question = st.text_input("คำถามลับ", value=config["secret_question"])
        new_answer   = st.text_input("คำตอบ", value=config["secret_answer"])
        if st.form_submit_button("💾 บันทึกคำถามลับ"):
            if not new_question or not new_answer:
                st.error("กรุณากรอกให้ครบ")
            else:
                config["secret_question"] = new_question
                config["secret_answer"]   = new_answer
                save_config(config)
                st.success("✅ บันทึกคำถามลับเรียบร้อย!")
